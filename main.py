# run  "pip install openpyxl" before running this program
#This program works for any amount of students, just make sure to follow the format as seen in the template gradebook xlsx file

from openpyxl import load_workbook
wb = load_workbook('Gradebook.xlsx')
sheet = wb.active

#Building the Variables for a huge for loop call at the end:
number_of_students = sheet.max_row - 1 #Because the first row is all headers
x = 0
lst = [[],[],[],[],[],[]]
z=1
for x in range(0,6): #Im going to use a nested list because i dont want to repeat code by iterating like this for every grade list
    
    for i in range(1, number_of_students+1):
        cell_obj = sheet.cell(row = i + 1, column = z) #Not tryna have 'Names:' as one of my students
        lst[x].append(cell_obj.value)
    z+=1
students = lst[0]
Q1_Grades = lst[1]
Q2_Grades = lst[2]
Q3_Grades = lst[3]
Q4_Grades = lst[4]
Desired_Grades_List = lst[5]



class Desired_Grade_Calculator:
    #Init runs without being called i think
    def __init__(self, name, Q1, Q2, Q3, Q4, desired_grade, lowest_final):
        self.Q1 = Q1
        self.Q2 = Q2
        self.Q3 = Q3
        self.Q4 = Q4
        self.desired_grade = desired_grade
        self.name = name
        self.lowest_final = lowest_final
    def calculator(self, name, Q1, Q2, Q3, Q4, desired_grade, lowest_final):
        global student_index
        if self.Q4 == 'E' and self.Q3 == 'E':
            sheet.cell(row = student_index, column= 7).value = "Doesn't matter"
            sheet.cell(row = student_index, column= 8).value = 'E'
            return "Sorry {}, because you failed the last 2 quarters, you automatically fail the course.".format(self.name)
        #REMINDER: YOU CAN ONLY RETURN ONCE
        # THIS USES THE HCPSS 8020 GRADING POLICY CODE WILL NOT WORK WITH OTHER POLICIES
        quarter_grade_scale = {"A" : .9, "B": .675, "C" : .450, "D": .225, "E" : 0}
        grade_scale = {"A" : 3.5, "B": 2.5, "C" : 1.5, "D": .75, "E" : .749}
        total = 0
        response = ''
        while True:
            
            self.desired_grade = desired_grade
            self.name = name
            self.lowest_final = lowest_final
            student_index = students.index(self.name) + 2 # +2 to account for the fact that indexes start at 0 and we need this for row numbers which have a header we are ignoring.
            total = round(total + quarter_grade_scale[Q1] + quarter_grade_scale[Q2] + quarter_grade_scale[Q3] + quarter_grade_scale[Q4], 3) #Round so we dont get disgusting floating numbers
            if desired_grade == 'F':
                sheet.cell(row = student_index, column= 7).value = "Doesn't matter"
                sheet.cell(row = student_index, column= 8).value = 'E'
                return "Sorry {}, not only is your desired grade not possible, there's actually no possibility that you can pass this class.".format(self.name)
            if total >= grade_scale[desired_grade]:
                lowest_final = 'E'
                sheet.cell(row = student_index, column= 7).value = lowest_final
                sheet.cell(row = student_index, column= 8).value = desired_grade
                return response + "Congrats, {} you can bomb the final if you want a(n) {}!".format(self.name, self.desired_grade)
            if grade_scale[desired_grade]-total>.4:
                make_letter_up_1=ord(desired_grade)
                desired_grade = chr(make_letter_up_1+1) #Moves up the letter up one in the ASCII characters so it's a quick fix without external modules
                response = "Sorry {}, you can't acheive a(n) '{}' even with an 'A' on the final; ".format(self.name, self.desired_grade)
                continue
            if .3< grade_scale[desired_grade]-total <.4:
                lowest_final = 'A'
                
            if .2< grade_scale[desired_grade]-total <.3:
                lowest_final = 'B'
                
            if .1< grade_scale[desired_grade]-total <.2:
                lowest_final = 'C'
                
            if .0< grade_scale[desired_grade]-total <.1:
                lowest_final = 'D'
            self.desired_grade = desired_grade
            self.name = name
            self.lowest_final = lowest_final
            sheet.cell(row = student_index, column= 7).value = lowest_final
            sheet.cell(row = student_index, column= 8).value = desired_grade
            wb.save('Gradebook.xlsx')
            return response + "{}, you need to get at least a '{}' on the final to achieve a(n) '{}'".format(self.name, self.lowest_final, self.desired_grade)


#huge call
for i in range(0, number_of_students):
    student = Desired_Grade_Calculator(students[i], Q1_Grades[i], Q2_Grades[i], Q3_Grades[i], Q4_Grades[i], Desired_Grades_List[i], '?',)
    print(Desired_Grade_Calculator.calculator(student, students[i], Q1_Grades[i], Q2_Grades[i], Q3_Grades[i], Q4_Grades[i], Desired_Grades_List[i], '?'))
    #Two question marks because the lowest possible final and overall grade need to be calculated.

wb.save('Gradebook.xlsx')
